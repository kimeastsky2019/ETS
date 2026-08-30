"""문서 4채널 분해 — 글 · 표 · 그림 · 엑셀.

기존 `app._extract_text()` 는 PDF 를 PyPDF2 로 통짜 텍스트로 만들어 업로드한다.
에너지진단 보고서에서 그렇게 하면 **표의 행 구조가 78% 깨진다** (실측치).
숫자는 텍스트 스트림 어딘가에 남지만 "어느 행 어느 열의 값인가" 가 사라져서,
`18대 | 22kW | 25.7kW | 7,200h | 80% | 2,664,576kWh` 가 의미를 잃는다.

에너지진단 보고서는 숫자가 전부 표에 있다. 그래서 채널을 나눈다.

| 채널 | 무엇 | 앵커 | 쓰임 |
|---|---|---|---|
| text  | 문단·항목 서술 | page | 서술 검색, 문제점/개선방안 |
| table | 셀 그리드 | page + table_idx + (row, col) | **모든 수치의 출처** |
| image | 사진·도면·차트·로고 | page + image_idx | 증적, 명판 OCR(v0.2) |
| excel | 표 전체를 시트로 | sheet = page-table | 감리 제출, 재계산 |

앵커는 `(문서해시, 페이지, 표인덱스, 행, 열)` 이다. bbox 는 속성으로만 들고
ID 에 넣지 않는다 — 파서를 고치면 좌표가 밀려 그래프가 통째로 끊기기 때문이다.
(에너지진단 온톨로지 원칙 2)
"""

from __future__ import annotations

import hashlib
import io
import os
import re
from dataclasses import dataclass, field, asdict
from typing import Any

try:
    import pdfplumber
except ImportError:  # pragma: no cover
    pdfplumber = None


NUMERIC = re.compile(r"-?\d[\d,]*\.?\d*")
# 로고·머리말 같은 잡동사니 이미지를 걸러내는 최소 크기(px)
MIN_IMAGE_PX = 80


@dataclass
class TextBlock:
    page: int
    idx: int
    text: str
    char_len: int = 0

    def __post_init__(self):
        self.char_len = len(self.text)

    @property
    def anchor(self) -> str:
        return f"p{self.page}/t{self.idx}"


@dataclass
class TableBlock:
    page: int
    idx: int
    header: list[str]
    rows: list[list[str]]
    n_numeric_cells: int = 0
    caption: str = ""

    @property
    def anchor(self) -> str:
        return f"p{self.page}/tbl{self.idx}"

    @property
    def shape(self) -> tuple[int, int]:
        return (len(self.rows), len(self.header) if self.header else 0)

    def cell_anchor(self, row: int, col: int) -> str:
        return f"{self.anchor}/r{row}c{col}"


@dataclass
class ImageBlock:
    page: int
    idx: int
    width: int
    height: int
    kind: str = "photo"          # photo | drawing | chart | logo | unknown
    nearby_caption: str = ""

    @property
    def anchor(self) -> str:
        return f"p{self.page}/img{self.idx}"


@dataclass
class ParsedDocument:
    filename: str
    doc_hash: str
    n_pages: int
    text_blocks: list[TextBlock] = field(default_factory=list)
    tables: list[TableBlock] = field(default_factory=list)
    images: list[ImageBlock] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)

    # ---- 파생 통계 -------------------------------------------------------
    @property
    def full_text(self) -> str:
        return "\n".join(b.text for b in self.text_blocks)

    @property
    def n_numeric_cells(self) -> int:
        return sum(t.n_numeric_cells for t in self.tables)

    def summary(self) -> dict[str, Any]:
        return {
            "filename": self.filename,
            "doc_hash": self.doc_hash,
            "pages": self.n_pages,
            "text_blocks": len(self.text_blocks),
            "text_chars": sum(b.char_len for b in self.text_blocks),
            "tables": len(self.tables),
            "table_rows": sum(len(t.rows) for t in self.tables),
            "numeric_cells": self.n_numeric_cells,
            "images": len(self.images),
            "image_kinds": _count(i.kind for i in self.images),
            "warnings": self.warnings,
        }


def _count(it) -> dict[str, int]:
    out: dict[str, int] = {}
    for v in it:
        out[v] = out.get(v, 0) + 1
    return out


def _clean(cell: Any) -> str:
    if cell is None:
        return ""
    return re.sub(r"\s+", " ", str(cell)).strip()


def _is_numeric(s: str) -> bool:
    if not s:
        return False
    m = NUMERIC.fullmatch(s.replace(" ", ""))
    return bool(m)


def _classify_image(width: int, height: int, page_text: str) -> str:
    """이미지 종류 추정. 확정이 아니라 제안 — 사람이 뒤집을 수 있다."""
    if width < MIN_IMAGE_PX or height < MIN_IMAGE_PX:
        return "logo"
    ratio = width / max(height, 1)
    txt = page_text or ""
    if any(k in txt for k in ("차트", "추이", "그래프")):
        return "chart"
    if any(k in txt for k in ("도면", "배치도", "구조도", "외형도", "제작도")):
        return "drawing"
    if ratio > 3.0 or ratio < 0.33:
        return "drawing"
    return "photo"


def parse_pdf(path: str, *, extract_images: bool = True) -> ParsedDocument:
    """PDF 를 4채널로 분해한다.

    표 추출은 pdfplumber 의 격자 인식을 쓴다. 격자가 없는 표(선 없는 레이아웃)는
    놓칠 수 있고, 그 경우 경고를 남긴다 — 조용히 빈 결과를 주지 않는다.
    """
    if pdfplumber is None:
        raise RuntimeError("pdfplumber 가 필요합니다: pip install pdfplumber")

    with open(path, "rb") as f:
        raw = f.read()
    doc_hash = hashlib.sha256(raw).hexdigest()[:16]

    doc = ParsedDocument(
        filename=os.path.basename(path), doc_hash=doc_hash, n_pages=0
    )

    with pdfplumber.open(io.BytesIO(raw)) as pdf:
        doc.n_pages = len(pdf.pages)
        for pno, page in enumerate(pdf.pages, start=1):
            page_text = page.extract_text() or ""

            # --- 글 -------------------------------------------------------
            if page_text.strip():
                for bidx, chunk in enumerate(_split_paragraphs(page_text)):
                    doc.text_blocks.append(TextBlock(page=pno, idx=bidx, text=chunk))

            # --- 표 -------------------------------------------------------
            try:
                raw_tables = page.extract_tables()
            except Exception as e:  # pragma: no cover
                raw_tables = []
                doc.warnings.append(f"p{pno} 표 추출 실패: {e}")

            tidx = 0
            for t in raw_tables:
                grid = [[_clean(c) for c in row] for row in t]
                grid = [r for r in grid if any(c for c in r)]
                if len(grid) < 2:
                    continue  # 사진 캡션용 1행짜리 격자는 표가 아니다
                header, rows = grid[0], grid[1:]
                n_num = sum(1 for r in rows for c in r if _is_numeric(c))
                if n_num == 0 and len(rows) < 2:
                    continue
                doc.tables.append(
                    TableBlock(
                        page=pno, idx=tidx, header=header, rows=rows,
                        n_numeric_cells=n_num,
                        caption=_guess_caption(page_text),
                    )
                )
                tidx += 1

            # --- 그림 -----------------------------------------------------
            if extract_images:
                for iidx, im in enumerate(page.images or []):
                    w = int(abs(im.get("x1", 0) - im.get("x0", 0)))
                    h = int(abs(im.get("bottom", 0) - im.get("top", 0)))
                    doc.images.append(
                        ImageBlock(
                            page=pno, idx=iidx, width=w, height=h,
                            kind=_classify_image(w, h, page_text),
                            nearby_caption=_guess_caption(page_text),
                        )
                    )

    if not doc.tables:
        doc.warnings.append(
            "표가 하나도 추출되지 않았습니다. 격자선 없는 레이아웃이거나 스캔본일 수 있습니다 "
            "(v0.2 OCR 대상)."
        )
    return doc


def _split_paragraphs(text: str) -> list[str]:
    parts = [p.strip() for p in re.split(r"\n\s*\n", text) if p.strip()]
    return parts or ([text.strip()] if text.strip() else [])


CAPTION = re.compile(r"\[\s*([^\]]{2,40})\s*\]")


def _guess_caption(page_text: str) -> str:
    m = CAPTION.search(page_text or "")
    return m.group(1).strip() if m else ""


# --------------------------------------------------------------------------
# 엑셀 채널 — 표를 그대로 시트로 떨군다. LLMWiki 의 Excel 내보내기와 같은 역할.
# --------------------------------------------------------------------------
def to_excel(doc: ParsedDocument, out_path: str) -> str:
    """추출한 표 전체를 xlsx 로. 시트 1장 = 표 1개, 첫 시트는 목차."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError as e:  # pragma: no cover
        raise RuntimeError("openpyxl 이 필요합니다: pip install openpyxl") from e

    wb = Workbook()
    idx_ws = wb.active
    idx_ws.title = "목차"
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="4F46E5")

    idx_ws.append(["시트", "페이지", "표", "행", "열", "숫자셀", "캡션"])
    for c in idx_ws[1]:
        c.font, c.fill = hdr_font, hdr_fill

    for i, t in enumerate(doc.tables, start=1):
        sheet = f"p{t.page}_t{t.idx}"[:31]
        ws = wb.create_sheet(sheet)
        if t.header:
            ws.append(t.header)
            for c in ws[1]:
                c.font, c.fill = hdr_font, hdr_fill
                c.alignment = Alignment(horizontal="center", wrap_text=True)
        for row in t.rows:
            ws.append(row)
        for col in ws.columns:
            width = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max(width + 2, 8), 40)
        r, c = t.shape
        idx_ws.append([sheet, t.page, t.idx, r, c, t.n_numeric_cells, t.caption])

    for col in idx_ws.columns:
        width = max((len(str(x.value or "")) for x in col), default=10)
        idx_ws.column_dimensions[col[0].column_letter].width = min(max(width + 2, 10), 40)

    wb.save(out_path)
    return out_path


# --------------------------------------------------------------------------
# RAG 인덱싱용 청크 — 채널마다 다르게 만든다.
# --------------------------------------------------------------------------
def to_chunks(doc: ParsedDocument) -> list[dict]:
    """검색 인덱스에 넣을 청크. **표는 표 단위로 통째 유지한다.**

    표를 문장처럼 잘라 넣으면 행-열 관계가 다시 깨진다. 표 1개 = 청크 1개로
    두고 마크다운 파이프 형식으로 직렬화해, 검색된 컨텍스트 안에서도 격자가
    살아 있게 한다.
    """
    chunks: list[dict] = []

    for b in doc.text_blocks:
        if b.char_len < 20:
            continue
        chunks.append({
            "channel": "text",
            "anchor": b.anchor,
            "page": b.page,
            "content": b.text,
        })

    for t in doc.tables:
        lines = []
        if t.caption:
            lines.append(f"[표: {t.caption}] (p.{t.page})")
        else:
            lines.append(f"[표] (p.{t.page})")
        if t.header:
            lines.append(" | ".join(t.header))
            lines.append(" | ".join("---" for _ in t.header))
        for row in t.rows:
            lines.append(" | ".join(row))
        chunks.append({
            "channel": "table",
            "anchor": t.anchor,
            "page": t.page,
            "content": "\n".join(lines),
            "n_numeric_cells": t.n_numeric_cells,
        })

    for im in doc.images:
        if im.kind == "logo":
            continue
        cap = im.nearby_caption or "캡션 없음"
        chunks.append({
            "channel": "image",
            "anchor": im.anchor,
            "page": im.page,
            "content": f"[그림/{im.kind}] p.{im.page} — {cap} ({im.width}x{im.height})",
        })

    return chunks


def to_dict(doc: ParsedDocument) -> dict:
    return {
        "summary": doc.summary(),
        "text_blocks": [asdict(b) for b in doc.text_blocks],
        "tables": [asdict(t) for t in doc.tables],
        "images": [asdict(i) for i in doc.images],
    }
